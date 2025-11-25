"""Report generation services for PDF and Excel output.

Generates formatted financial reports in multiple formats.
"""

from io import BytesIO, StringIO
from datetime import date
from typing import List, Dict, Optional
from decimal import Decimal
import csv
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generates reports in various formats."""

    def __init__(self):
        """Initialize report generator."""
        pass

    def generate_pdf(
        self,
        report_type: str,
        title: str,
        period_start: date,
        period_end: date,
        data: Dict,
        company_name: str = "Company",
        include_summary: bool = True
    ) -> bytes:
        """
        Generate PDF report.

        Args:
            report_type: Type of report (profit_loss, balance_sheet, etc)
            title: Report title
            period_start: Start date
            period_end: End date
            data: Report data (tables, sections)
            company_name: Organization name
            include_summary: Include executive summary

        Returns:
            PDF bytes
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = []

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1f4788'),
                spaceAfter=6,
                alignment=TA_CENTER
            )

            # Header
            story.append(Paragraph(company_name, title_style))
            story.append(Paragraph(title, styles['Heading2']))
            story.append(Paragraph(
                f"Period: {period_start.strftime('%b %d, %Y')} - {period_end.strftime('%b %d, %Y')}",
                styles['Normal']
            ))
            story.append(Spacer(1, 0.3*inch))

            # Summary section
            if include_summary and 'summary' in data:
                story.append(Paragraph("Executive Summary", styles['Heading3']))
                for key, value in data['summary'].items():
                    story.append(Paragraph(f"<b>{key}:</b> {value}", styles['Normal']))
                story.append(Spacer(1, 0.2*inch))

            # Data tables
            if 'tables' in data:
                for section_title, table_data in data['tables'].items():
                    story.append(Paragraph(section_title, styles['Heading3']))
                    if table_data:
                        table = Table(table_data, colWidths=[2.5*inch, 1.2*inch])
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 11),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                            ('FONTSIZE', (0, 1), (-1, -1), 9),
                        ]))
                        story.append(table)
                    story.append(Spacer(1, 0.2*inch))

            # Footer
            story.append(Spacer(1, 0.3*inch))
            story.append(Paragraph(
                f"<i>Generated on {date.today().strftime('%B %d, %Y')}</i>",
                styles['Normal']
            ))

            doc.build(story)
            buffer.seek(0)
            logger.info(f"Generated PDF report: {report_type}")
            return buffer.getvalue()

        except ImportError:
            logger.error("reportlab not installed, using text fallback")
            return self._generate_text_report(report_type, title, period_start, period_end, data)

    def generate_excel(
        self,
        report_type: str,
        title: str,
        period_start: date,
        period_end: date,
        data: Dict,
        company_name: str = "Company"
    ) -> bytes:
        """
        Generate Excel report.

        Args:
            report_type: Type of report
            title: Report title
            period_start: Start date
            period_end: End date
            data: Report data
            company_name: Organization name

        Returns:
            Excel bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = report_type[:31]  # Sheet name max 31 chars

            # Header styling
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            row = 1
            ws[f'A{row}'] = company_name
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1

            ws[f'A{row}'] = title
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            ws[f'A{row}'] = f"Period: {period_start} to {period_end}"
            row += 2

            # Summary section
            if 'summary' in data:
                ws[f'A{row}'] = "Executive Summary"
                ws[f'A{row}'].font = Font(bold=True, size=11)
                row += 1

                for key, value in data['summary'].items():
                    ws[f'A{row}'] = key
                    ws[f'B{row}'] = str(value)
                    row += 1
                row += 1

            # Data tables
            if 'tables' in data:
                for section_title, table_data in data['tables'].items():
                    ws[f'A{row}'] = section_title
                    ws[f'A{row}'].font = Font(bold=True, size=11)
                    row += 1

                    if table_data:
                        for col_idx, cell_value in enumerate(table_data[0], 1):
                            cell = ws.cell(row=row, column=col_idx, value=cell_value)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.border = border
                            cell.alignment = Alignment(horizontal='center')
                        row += 1

                        for record in table_data[1:]:
                            for col_idx, cell_value in enumerate(record, 1):
                                cell = ws.cell(row=row, column=col_idx, value=cell_value)
                                cell.border = border
                                if col_idx > 1:
                                    cell.alignment = Alignment(horizontal='right')
                            row += 1
                    row += 1

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            logger.info(f"Generated Excel report: {report_type}")
            return buffer.getvalue()

        except ImportError:
            logger.error("openpyxl not installed, using CSV fallback")
            return self.generate_csv(report_type, title, period_start, period_end, data).encode('utf-8')

    def generate_csv(
        self,
        report_type: str,
        title: str,
        period_start: date,
        period_end: date,
        data: Dict
    ) -> str:
        """
        Generate CSV report.

        Args:
            report_type: Type of report
            title: Report title
            period_start: Start date
            period_end: End date
            data: Report data

        Returns:
            CSV string
        """
        buffer = StringIO()
        writer = csv.writer(buffer)

        writer.writerow([title])
        writer.writerow([f"Period: {period_start} to {period_end}"])
        writer.writerow([])

        # Summary
        if 'summary' in data:
            writer.writerow(["Executive Summary"])
            for key, value in data['summary'].items():
                writer.writerow([key, value])
            writer.writerow([])

        # Tables
        if 'tables' in data:
            for section_title, table_data in data['tables'].items():
                writer.writerow([section_title])
                if table_data:
                    writer.writerows(table_data)
                writer.writerow([])

        logger.info(f"Generated CSV report: {report_type}")
        return buffer.getvalue()

    def _generate_text_report(
        self,
        report_type: str,
        title: str,
        period_start: date,
        period_end: date,
        data: Dict
    ) -> bytes:
        """Generate plain text report as fallback."""
        lines = [
            title,
            f"Period: {period_start} to {period_end}",
            "",
        ]

        if 'summary' in data:
            lines.append("Executive Summary")
            for key, value in data['summary'].items():
                lines.append(f"  {key}: {value}")
            lines.append("")

        if 'tables' in data:
            for section_title, table_data in data['tables'].items():
                lines.append(section_title)
                if table_data:
                    lines.extend([str(row) for row in table_data])
                lines.append("")

        logger.info(f"Generated text report: {report_type}")
        return "\n".join(lines).encode('utf-8')
